import openpyxl
import sqlite3
from sqlite3 import Error

path = '../data/'
majors = ['Biological Sciences', 'Business', 'Communication Arts',
            'Computer Information Systems (1)', 'Computer Science (1)',
            'EET w_ CT Option (1)', 'Electrical Engineering Technology (1)',
            'English (1)', 'English Teaching (1)', 'History (1)',
            'Mechanical Engineering Technology (1)', 'Politics and Society (1)',
            'Psychology (1)', 'Sign Language Interpretation (1)'
        ]
ext = '.xlsx'
db_file = '../../db.sqlite3'



def addSheetToList(sheet, school_list):
    for row_idx in range(sheet.min_row+1, sheet.max_row + 1):
        cell_content = sheet.cell(row=row_idx, column=1).value
        if cell_content not in school_list:
            if cell_content != None:
                    school_list.append(cell_content)
    return school_list


def create_connection(db_file):
    conn = None
    try:
        conn = sqlite3.connect(db_file)
    except Error as e:
        print(e)
    return conn

def create_school(conn, school):
    sql = ''' INSERT INTO transfer_school(school_name) VALUES(?) '''
    cur = conn.cursor()
    cur.execute(sql, [school])
    return cur.lastrowid

if __name__ == '__main__':
    school_list = []
    for m in majors:
        file = path + m + ext
        wb = openpyxl.load_workbook(file)
        sheet = wb.active
        school_list = addSheetToList(sheet, school_list)
    print(school_list)
    conn = create_connection(db_file)
    with conn:
        for school in school_list:
            task = create_school(conn, school)
