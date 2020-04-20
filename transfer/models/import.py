import openpyxl
import sqlite3
from sqlite3 import Error

data_file = '../data/transfer_by_major.xlsx'
db_file = '../../db.sqlite3'

wb = openpyxl.load_workbook(data_file)
majors = wb.sheetnames

def addSheetToList(sheet, school_list):
    for row_idx in range(sheet.min_row+1, sheet.max_row + 1):
        cell_content = sheet.cell(row=row_idx, column=1).value
        if cell_content not in school_list:
            if cell_content != None:
                    school_list.append(cell_content)
    return school_list


def create_connection(db_file):
    conn = None
    try:
        conn = sqlite3.connect(db_file)
    except Error as e:
        print(e)
    return conn

def create_school_list(school_list, sheet):
    school_list = addSheetToList(sheet, school_list)
    return school_list



def create_major(conn, major):
    sql = ''' INSERT INTO transfer_major(major_name) VALUES(?) '''
    cur = conn.cursor()
    cur.execute(sql, [major])
    return cur.lastrowid

def create_school(conn, school, major_id):
    sql = ''' INSERT INTO transfer_school(school_name, major_id_id) VALUES(?,?) '''
    cur = conn.cursor()
    cur.execute(sql, [school, major_id])
    return cur.lastrowid

if __name__ == '__main__':
    conn = create_connection(db_file)
    with conn:
        for major in majors:
            task = create_major(conn, major)
    for i in range(1,15):
        school_list = []
        sheet = wb[majors[i]]
        school_list = create_school_list(school_list, sheet)
        for school in school_list:
            with conn:
                task = create_school(conn, school, i)
